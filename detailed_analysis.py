import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def analyze_data_tables(ws):
    """Extract actual data tables from worksheets"""
    tables = []
    current_table = None

    for row_idx in range(1, min(ws.max_row + 1, 200)):
        row_data = []
        has_data = False

        for col_idx in range(1, min(ws.max_column + 1, 50)):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue

            if cell.value is not None:
                has_data = True
                row_data.append({
                    'col': get_column_letter(col_idx),
                    'value': str(cell.value)[:100],
                    'type': cell.data_type,
                    'format': cell.number_format
                })

        if has_data and row_data:
            if current_table is None:
                current_table = {'start_row': row_idx, 'rows': []}
            current_table['rows'].append({'row_num': row_idx, 'data': row_data})
        elif current_table is not None:
            current_table['end_row'] = row_idx - 1
            if len(current_table['rows']) > 2:  # Only save meaningful tables
                tables.append(current_table)
            current_table = None

    return tables

def main():
    file_path = r"C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Blank OECS MS Template.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=False)

    report = []
    report.append("="*100)
    report.append("COMPREHENSIVE OECS EDUCATION DATA TEMPLATE ANALYSIS")
    report.append("="*100)
    report.append("")

    # Overall summary
    report.append(f"FILE: {file_path}")
    report.append(f"TOTAL WORKSHEETS: {len(wb.sheetnames)}")
    report.append(f"ACADEMIC YEAR: 2023-2024")
    report.append("")

    # Detailed sheet analysis
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        report.append("")
        report.append("="*100)
        report.append(f"SHEET: {sheet_name}")
        report.append("="*100)
        report.append("")

        # Get year from first row if available
        year_cell = None
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value and '2023' in str(cell.value) or '2024' in str(cell.value):
                    year_cell = cell.value
                    break
            if year_cell:
                break

        if year_cell:
            report.append(f"Academic Year: {year_cell}")

        # Document structure
        report.append(f"Dimensions: {ws.dimensions}")
        report.append(f"Data Range: Rows 1-{ws.max_row}, Columns A-{get_column_letter(ws.max_column)}")
        report.append(f"Merged Cells: {len(ws.merged_cells.ranges)}")
        report.append("")

        # Analyze tables
        tables = analyze_data_tables(ws)
        report.append(f"NUMBER OF DATA TABLES IDENTIFIED: {len(tables)}")
        report.append("")

        # Document each table
        for i, table in enumerate(tables, 1):
            report.append(f"  TABLE {i}:")
            report.append(f"    Row Range: {table['start_row']}-{table.get('end_row', 'ongoing')}")
            report.append(f"    Total Rows: {len(table['rows'])}")

            # Show first few rows as structure
            report.append("    Structure (first 5 rows):")
            for row in table['rows'][:5]:
                row_str = f"      Row {row['row_num']}: "
                data_parts = []
                for cell in row['data'][:8]:  # Limit columns shown
                    val = cell['value'][:30]
                    data_parts.append(f"{cell['col']}={val}")
                report.append(row_str + " | ".join(data_parts))
            report.append("")

        # Document formulas
        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row)):
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formulas.append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

        if formulas:
            report.append(f"FORMULAS ({len(formulas)} found):")
            formula_patterns = {}
            for f in formulas[:20]:
                pattern = f['formula'].split('(')[0] if '(' in f['formula'] else f['formula']
                if pattern not in formula_patterns:
                    formula_patterns[pattern] = []
                formula_patterns[pattern].append(f"{f['cell']}: {f['formula']}")

            for pattern, examples in formula_patterns.items():
                report.append(f"  Pattern: {pattern}")
                for ex in examples[:3]:
                    report.append(f"    - {ex}")
            report.append("")

        # Document data validations
        if ws.data_validations.dataValidation:
            report.append(f"DATA VALIDATIONS ({len(ws.data_validations.dataValidation)} found):")
            for dv in ws.data_validations.dataValidation[:5]:
                report.append(f"  Type: {dv.type}")
                report.append(f"  Range: {dv.sqref}")
                if dv.formula1:
                    report.append(f"  Allowed Values: {dv.formula1}")
                report.append("")

        # Document comments
        comments = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append({
                        'cell': cell.coordinate,
                        'comment': cell.comment.text
                    })

        if comments:
            report.append(f"COMMENTS ({len(comments)} found):")
            for c in comments[:3]:
                report.append(f"  {c['cell']}: {c['comment'][:100]}")
            report.append("")

    # Save report
    output_file = r"C:\Users\Clendon\oecs-education-dashboard\comprehensive_template_report.txt"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report))

    print('\n'.join(report))
    print(f"\n\nReport saved to: {output_file}")

if __name__ == "__main__":
    main()
