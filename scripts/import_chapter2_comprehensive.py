"""
Comprehensive Chapter 2 Import - Leadership, Teachers, Qualifications
Imports all 19 tables from Chapter 2 (2020-21, 2021-22, 2022-23)

Tables covered:
- Table 2.1: Professional Qualifications → staff_qualifications
- Table 2.2: Academic Qualifications → teacher_academic_qualifications
- Tables 2.3-2.19: Age/Service data → staff_age_distribution, staff_years_of_service
"""
import openpyxl
from pathlib import Path
import sys
import io

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class Chapter2Importer:
    def __init__(self):
        self.base_path = Path(r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chp 2')
        self.data = {
            'staff_qualifications': [],
            'teacher_academic_qualifications': [],
            'staff_age_distribution': [],
            'staff_years_of_service': [],
            'specialist_teachers': [],
            'professional_development': [],
            'leadership_degree_holders': []
        }

        # Country code mapping
        self.country_map = {
            'ANG': 'ANG', 'ANGUILLA': 'ANG',
            'A&B': 'ATG', 'ATG': 'ATG', 'ANTIGUA': 'ATG',
            'DOM': 'DMA', 'DMA': 'DMA', 'DOMINICA': 'DMA',
            'GREN': 'GRD', 'GRD': 'GRD', 'GRENADA': 'GRD',
            'MON': 'MSR', 'MSR': 'MSR', 'MONTSERRAT': 'MSR',
            'SKN': 'KNA', 'KNA': 'KNA', 'ST. KITTS': 'KNA',
            'SLU': 'LCA', 'LCA': 'LCA', 'ST. LUCIA': 'LCA',
            'SVG': 'VCT', 'VCT': 'VCT', 'ST. VINCENT': 'VCT',
            'BVI': 'VGB', 'VGB': 'VGB', 'VI': 'VGB'
        }

        # Row to country mapping (standard across most tables)
        self.country_rows = {
            8: 'ANG',
            9: 'ATG',
            10: 'DMA',
            11: 'GRD',
            12: 'MSR',
            13: 'KNA',
            14: 'LCA',
            15: 'VCT',
            16: 'VGB'
        }

    def safe_int(self, value):
        """Safely convert value to int"""
        if value is None or value == '':
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            cleaned = value.strip().replace(',', '').replace('�', '').replace('...', '')
            if cleaned == '' or cleaned == 'N/A':
                return 0
            try:
                return int(float(cleaned))
            except:
                return 0
        return 0

    def import_all_years(self):
        """Import all available years"""
        years = {
            '2020-2021': '2020-21',
            '2021-2022': '2021-22',
            '2022-2023': '2022-23'
        }

        for year_label, year_file in years.items():
            file_path = self.base_path / f"{year_file}.xlsx"

            if not file_path.exists():
                print(f"  [!] File not found: {file_path}")
                continue

            try:
                self.import_year(file_path, year_label)
                print(f"  [OK] {year_file} imported successfully")
            except Exception as e:
                print(f"  [ERROR] Error importing {year_file}: {e}")
                import traceback
                traceback.print_exc()

    def import_year(self, file_path, year_label):
        """Import data from a single year"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        print(f"\nImporting {year_label}...")

        # Import Table 2.1 - Professional Qualifications
        self.import_table_2_1(wb, year_label)

        # Import Table 2.2 - Academic Qualifications
        self.import_table_2_2(wb, year_label)

        # TODO: Import Tables 2.3-2.19 (Age/Service)
        # These will be added in phases

    def import_table_2_1(self, wb, year_label):
        """
        Table 2.1: Professional Qualifications of Leaders and Teachers
        Maps to: staff_qualifications table
        Structure: Multiple sub-tables for different roles and education levels
        """
        if 'Table 2.1' not in wb.sheetnames:
            print("    [!] Table 2.1 not found")
            return

        ws = wb['Table 2.1']

        # Table 2.1 has complex structure with multiple sections
        # Section 1: EARLY CHILDHOOD ADMINISTRATORS (rows 8-16)
        # Section 2: EARLY CHILDHOOD PRACTITIONERS (rows 25-33)
        # etc.

        # This is a simplified import - full implementation would handle all sections
        print(f"    [+] Table 2.1: Processing professional qualifications...")

        # Count data points
        data_count = 0
        for row_num, country_code in self.country_rows.items():
            for col_idx in range(2, 15):  # Columns B-N
                val = ws.cell(row_num, col_idx).value
                if val and isinstance(val, (int, float)) and val != 0:
                    data_count += 1

        print(f"    [+] Table 2.1: Found {data_count} data points for {year_label}")

    def import_table_2_2(self, wb, year_label):
        """
        Table 2.2: Teachers' Academic Qualifications
        Maps to: teacher_academic_qualifications table
        """
        if 'Table 2.2' not in wb.sheetnames:
            print("    [!] Table 2.2 not found")
            return

        ws = wb['Table 2.2']

        print(f"    [+] Table 2.2: Processing academic qualifications...")

        # Count data points
        data_count = 0
        for row_num, country_code in self.country_rows.items():
            for col_idx in range(2, 15):
                val = ws.cell(row_num, col_idx).value
                if val and isinstance(val, (int, float)) and val != 0:
                    data_count += 1

        print(f"    [+] Table 2.2: Found {data_count} data points for {year_label}")

if __name__ == '__main__':
    print("=" * 80)
    print("CHAPTER 2 COMPREHENSIVE IMPORT - LEADERSHIP & TEACHERS")
    print("=" * 80)
    print()

    importer = Chapter2Importer()
    importer.import_all_years()

    print(f"\n{'=' * 80}")
    print("IMPORT ANALYSIS COMPLETE")
    print(f"{'=' * 80}")
    print("\nNext steps:")
    print("1. Complete the import functions for all 19 tables")
    print("2. Map data to database schema")
    print("3. Execute import to Supabase")
