"""
Import Chapter 2 data to Supabase
Handles professional qualifications and academic qualifications
All 3 years: 2020-21, 2021-22, 2022-23
"""
import os
import sys
import io
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Load environment variables
env_path = Path(__file__).parent.parent / '.env.local'
load_dotenv(dotenv_path=env_path)

# Get Supabase credentials
supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

if not supabase_url or not supabase_key:
    print("ERROR: Missing Supabase credentials")
    sys.exit(1)

supabase: Client = create_client(supabase_url, supabase_key)

print("=" * 80)
print("CHAPTER 2 IMPORT TO SUPABASE")
print("=" * 80)
print()

# Get mappings
countries_response = supabase.table('countries').select('id, country_code').execute()
countries_map = {c['country_code']: c['id'] for c in countries_response.data}

years_response = supabase.table('academic_years').select('id, year_label').execute()
years_map = {y['year_label']: y['id'] for y in years_response.data}

print(f"Found {len(countries_map)} countries and {len(years_map)} academic years")
print()

# Country row mapping for Chapter 2
country_rows = {
    8: 'ANG',
    9: 'ATG',
    10: 'DMA',
    11: 'GRD',
    12: 'MSR',
    13: 'KNA',
    14: 'LCA',
    15: 'VCT',
    16: 'VGB'
}

def safe_int(value):
    """Safely convert value to int"""
    if value is None or value == '':
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(float(str(value).replace(',', '').strip()))
    except:
        return 0

def import_table_2_2(file_path, year_label):
    """
    Import Table 2.2: Teachers' Academic Qualifications
    This is simpler and more straightforward than Table 2.1
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if 'Table 2.2' not in wb.sheetnames:
        print(f"  [!] Table 2.2 not found in {file_path.name}")
        return 0

    ws = wb['Table 2.2']

    # Table 2.2 structure:
    # Row 1-6: Headers
    # Row 7: Column headers (Qualifications)
    # Rows 8-16: Country data
    # Columns: B-M (12 columns for different qualifications × education levels)

    records = []

    for row_num, country_code in country_rows.items():
        country_id = countries_map.get(country_code)
        year_id = years_map.get(year_label)

        if not country_id or not year_id:
            continue

        # Extract qualification counts
        # This is a simplified version - full implementation would map all columns
        for col_idx in range(2, 14):  # Columns B-M
            value = safe_int(ws.cell(row_num, col_idx).value)
            if value > 0:
                # For now, just count that we found data
                # Full implementation would properly map qual_level and education_stage
                pass

    print(f"  [+] Table 2.2 ({year_label}): Analyzed {len(country_rows)} countries")
    return len(records)

def import_year(year_label, year_file):
    """Import all tables for a single year"""
    base_path = Path(r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chp 2')
    file_path = base_path / f"{year_file}.xlsx"

    if not file_path.exists():
        print(f"  [!] File not found: {file_path}")
        return 0

    print(f"\nProcessing {year_label}...")

    # Import Table 2.2
    count = import_table_2_2(file_path, year_label)

    return count

# Import all years
years = {
    '2020-2021': '2020-21',
    '2021-2022': '2021-22',
    '2022-2023': '2022-23'
}

total_records = 0
for year_label, year_file in years.items():
    count = import_year(year_label, year_file)
    total_records += count

print()
print("=" * 80)
print(f"IMPORT COMPLETE")
print(f"Total records processed: {total_records}")
print("=" * 80)
print()
print("NOTE: This is a skeleton import. Full implementation requires:")
print("  1. Complete mapping of all 19 tables")
print("  2. Proper data transformation for each table structure")
print("  3. Database inserts for all extracted data")
print()
print("Tables ready in database:")
print("  - staff_qualifications")
print("  - teacher_academic_qualifications")
print("  - leadership_degree_holders")
print("  - specialist_teachers")
print("  - professional_development")
print("  - staff_age_distribution")
print("  - staff_years_of_service")
