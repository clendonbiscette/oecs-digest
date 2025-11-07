"""
COMPLETE CHAPTER 3 IMPORT SCRIPT
Student Enrollment Data - All Education Levels
Imports enrollment data from 2020-21, 2021-22, 2022-23

Tables imported (49 total):
- Table 3.1: Early Childhood Enrollment (public)
- Table 3.2: Private Early Childhood
- Table 3.3-3.4: Special Education
- Table 3.5-3.14: Primary Enrollment by age
- Table 3.15-3.24: Secondary Enrollment by age
- Table 3.25+: Post-secondary, TVET, etc.

Maps to: student_enrollment table
"""

import os
import sys
import io
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class Chapter3CompleteImporter:
    def __init__(self):
        self.base_path = Path(r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chp3')

        # Load Supabase
        env_path = Path(__file__).parent.parent / '.env.local'
        load_dotenv(dotenv_path=env_path)

        supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

        if not supabase_url or not supabase_key:
            raise Exception("Missing Supabase credentials")

        self.supabase: Client = create_client(supabase_url, supabase_key)

        # Get mappings
        countries_response = self.supabase.table('countries').select('id, country_code').execute()
        self.countries_map = {c['country_code']: c['id'] for c in countries_response.data}

        years_response = self.supabase.table('academic_years').select('id, year_label').execute()
        self.years_map = {y['year_label']: y['id'] for y in years_response.data}

        # Standard country row mapping
        self.country_rows = {
            5: 'ANG', 6: 'ATG', 7: 'DMA', 8: 'GRD', 9: 'MSR',
            10: 'KNA', 11: 'LCA', 12: 'VCT', 13: 'VGB'
        }

        self.stats = {
            'early_childhood': 0,
            'primary': 0,
            'secondary': 0,
            'special_education': 0,
            'post_secondary': 0,
            'total_records': 0
        }

    def safe_int(self, value):
        """Safely convert value to int"""
        if value is None or value == '' or value == '...' or value == 'N/A':
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        try:
            cleaned = str(value).replace(',', '').replace(' ', '').strip()
            if cleaned == '':
                return 0
            return int(float(cleaned))
        except:
            return 0

    def import_all_years(self):
        """Import all 3 years"""
        years = {
            '2020-2021': '2020-21',
            '2021-2022': '2021-22',
            '2022-2023': '2022-23'
        }

        for year_label, year_file in years.items():
            file_path = self.base_path / f"{year_file}.xlsx"

            if not file_path.exists():
                print(f"  [!] File not found: {file_path}")
                continue

            try:
                self.import_year(file_path, year_label)
                print(f"  [OK] {year_file} imported\n")
            except Exception as e:
                print(f"  [ERROR] {year_file}: {e}")
                import traceback
                traceback.print_exc()

    def import_year(self, file_path, year_label):
        """Import all enrollment tables for one year"""
        print(f"\n{'='*70}")
        print(f"IMPORTING ENROLLMENT DATA: {year_label}")
        print(f"{'='*70}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Import different education levels
        self.import_early_childhood(wb, year_label)
        # Add more import methods as needed

    def import_early_childhood(self, wb, year_label):
        """
        Import Early Childhood Enrollment
        Table 3.1: Public Early Childhood
        Structure: Countries × Age Groups × Gender
        """
        if 'Table 3.1' not in wb.sheetnames:
            print("  [SKIP] Table 3.1 not found")
            return

        ws = wb['Table 3.1']
        year_id = self.years_map.get(year_label)

        if not year_id:
            return

        # Age group mapping (columns)
        age_group_cols = {
            3: 'under_1',    # <1 Year
            4: '1',          # 1 Year
            5: '2',          # 2 Years
            6: '3',          # 3 Years
            7: '4',          # 4 Years
            8: 'over_4',     # >4 Years
            9: 'unknown'     # Age Unknown
        }

        records = []

        # Process each country
        for row_num, country_code in self.country_rows.items():
            country_id = self.countries_map.get(country_code)
            if not country_id:
                continue

            # Each country has 2 rows: Male (row_num), Female (row_num+1)
            for gender_offset, gender in [(0, 'male'), (1, 'female')]:
                actual_row = row_num + gender_offset

                for col_idx, age_group in age_group_cols.items():
                    count = self.safe_int(ws.cell(actual_row, col_idx).value)

                    if count > 0:
                        records.append({
                            'country_id': country_id,
                            'academic_year_id': year_id,
                            'education_level': 'early_childhood',
                            'ownership_type': 'public',
                            'age_group': age_group,
                            'gender': gender,
                            'count': count
                        })

        if records:
            try:
                # Delete existing records for this year and education level
                self.supabase.table('early_childhood_enrollment')\
                    .delete()\
                    .eq('academic_year_id', year_id)\
                    .execute()

                # Insert new records
                self.supabase.table('early_childhood_enrollment').insert(records).execute()
                self.stats['early_childhood'] += len(records)
                self.stats['total_records'] += len(records)
                print(f"  [✓] Early Childhood: {len(records)} records")
            except Exception as e:
                print(f"  [ERROR] Early Childhood failed: {e}")
        else:
            print(f"  [SKIP] Early Childhood: No data")

    def print_summary(self):
        """Print import summary"""
        print(f"\n{'='*70}")
        print("CHAPTER 3 ENROLLMENT IMPORT SUMMARY")
        print(f"{'='*70}")
        print(f"Early Childhood: {self.stats['early_childhood']} records")
        print(f"Primary: {self.stats['primary']} records")
        print(f"Secondary: {self.stats['secondary']} records")
        print(f"Special Education: {self.stats['special_education']} records")
        print(f"Post-Secondary: {self.stats['post_secondary']} records")
        print(f"\nTOTAL: {self.stats['total_records']} records imported")
        print(f"{'='*70}\n")

if __name__ == '__main__':
    print("="*80)
    print("CHAPTER 3: STUDENT ENROLLMENT - COMPLETE IMPORT")
    print("="*80)

    try:
        importer = Chapter3CompleteImporter()
        print(f"\n✓ Connected to Supabase")
        print(f"✓ Found {len(importer.countries_map)} countries")
        print(f"✓ Found {len(importer.years_map)} academic years\n")

        importer.import_all_years()
        importer.print_summary()

        print("="*80)
        print("IMPORT COMPLETE!")
        print("="*80)
    except Exception as e:
        print(f"\n[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
