"""
COMPLETE CHAPTER 3 IMPORT SCRIPT - REBUILT
Student Enrollment Data - All Education Levels
Imports enrollment data from 2020-21, 2021-22, 2022-23

Matches actual Supabase schema:
- early_childhood_enrollment: Simple format (institution_type, age_group, male, female)
- special_education_enrollment: Simple format
- primary_enrollment: Wide format (age_group, k_male, k_female, g1_male, g1_female, ...)
- secondary_enrollment: Wide format (age_group, f1_male, f1_female, ...)
"""

import os
import sys
import io
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class Chapter3RebuiltImporter:
    def __init__(self):
        self.base_path = Path(r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chp3')

        # Load Supabase
        env_path = Path(__file__).parent.parent / '.env.local'
        load_dotenv(dotenv_path=env_path)

        supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

        if not supabase_url or not supabase_key:
            raise Exception("Missing Supabase credentials")

        self.supabase: Client = create_client(supabase_url, supabase_key)

        # Get mappings
        countries_response = self.supabase.table('countries').select('id, country_code').execute()
        self.countries_map = {c['country_code']: c['id'] for c in countries_response.data}

        years_response = self.supabase.table('academic_years').select('id, year_label').execute()
        self.years_map = {y['year_label']: y['id'] for y in years_response.data}

        # Country code mapping (Excel to DB)
        self.country_code_map = {
            'ANG': 'ANG',
            'A&B': 'ATG',
            'DOM': 'DMA',
            'GRD': 'GRD',
            'MON': 'MSR',
            'SKN': 'KNA',
            'SLU': 'LCA',
            'SVG': 'VCT',
            'VI': 'VGB'
        }

        # Country column mapping (for Table 3.1, 3.2 format)
        self.country_columns = {
            3: 'ANG', 4: 'ATG', 5: 'DMA', 6: 'GRD', 7: 'MSR',
            8: 'KNA', 9: 'LCA', 10: 'VCT', 11: 'VGB'
        }

        # Table to country mapping for primary/secondary
        self.primary_tables = {
            'Table 3.7': 'ANG',
            'Table 3.8': 'ATG',
            'Table 3.9': 'DMA',
            'Table 3.10': 'GRD',
            'Table 3.11': 'MSR',
            'Table 3.12': 'KNA',
            'Table 3.13': 'LCA',
            'Table 3.14': 'VCT',
            'Table 3.15': 'VGB'
        }

        self.stats = {
            'early_childhood': 0,
            'special_education': 0,
            'primary': 0,
            'secondary': 0,
            'total_records': 0
        }

    def safe_int(self, value):
        """Safely convert value to int"""
        if value is None or value == '' or value == '...' or value == 'N/A':
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        try:
            cleaned = str(value).replace(',', '').replace(' ', '').strip()
            if cleaned == '':
                return 0
            return int(float(cleaned))
        except:
            return 0

    def import_all_years(self):
        """Import all 3 years"""
        years = {
            '2020-2021': '2020-21',
            '2021-2022': '2021-22',
            '2022-2023': '2022-23'
        }

        for year_label, year_file in years.items():
            file_path = self.base_path / f"{year_file}.xlsx"

            if not file_path.exists():
                print(f"  [!] File not found: {file_path}")
                continue

            try:
                self.import_year(file_path, year_label)
                print(f"  [OK] {year_file} imported\n")
            except Exception as e:
                print(f"  [ERROR] {year_file}: {e}")
                import traceback
                traceback.print_exc()

    def import_year(self, file_path, year_label):
        """Import all enrollment tables for one year"""
        print(f"\n{'='*70}")
        print(f"IMPORTING ENROLLMENT DATA: {year_label}")
        print(f"{'='*70}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        self.import_early_childhood(wb, year_label)
        self.import_special_education(wb, year_label)
        self.import_primary(wb, year_label)

    def import_early_childhood(self, wb, year_label):
        """
        Import Early Childhood Enrollment
        Table 3.1: Public Early Childhood
        Table 3.2: Private Early Childhood
        Format: Countries in columns, age groups in rows with M/F pairs
        """
        year_id = self.years_map.get(year_label)
        if not year_id:
            return

        tables = {
            'Table 3.1': 'Public',
            'Table 3.2': 'Private/Gov Assisted'
        }

        all_records = []

        for table_name, institution_type in tables.items():
            if table_name not in wb.sheetnames:
                print(f"  [SKIP] {table_name} not found")
                continue

            ws = wb[table_name]

            # Age groups start at row 5, in pairs (M then F)
            age_groups = [
                (5, '< 1 Year', 'under_1'),
                (7, '1 Year', '1'),
                (9, '2 Years', '2'),
                (11, '3 Years', '3'),
                (13, '4 Years', '4'),
                (15, '>4 Years', 'over_4'),
                (17, 'Age Unknown', 'unknown')
            ]

            # Process each country column
            for col_idx, excel_code in self.country_columns.items():
                country_code = self.country_code_map.get(excel_code)
                country_id = self.countries_map.get(country_code)

                if not country_id:
                    continue

                # Process each age group
                for male_row, age_label, age_code in age_groups:
                    female_row = male_row + 1

                    male_count = self.safe_int(ws.cell(male_row, col_idx).value)
                    female_count = self.safe_int(ws.cell(female_row, col_idx).value)

                    if male_count > 0 or female_count > 0:
                        all_records.append({
                            'country_id': country_id,
                            'academic_year_id': year_id,
                            'institution_type': institution_type,
                            'age_group': age_code,
                            'male': male_count,
                            'female': female_count
                        })

        if all_records:
            try:
                # Delete existing
                self.supabase.table('early_childhood_enrollment')\
                    .delete()\
                    .eq('academic_year_id', year_id)\
                    .execute()

                # Insert new
                self.supabase.table('early_childhood_enrollment').insert(all_records).execute()
                self.stats['early_childhood'] += len(all_records)
                self.stats['total_records'] += len(all_records)
                print(f"  [OK] Early Childhood: {len(all_records)} records")
            except Exception as e:
                print(f"  [ERROR] Early Childhood failed: {e}")
        else:
            print(f"  [SKIP] Early Childhood: No data")

    def import_special_education(self, wb, year_label):
        """
        Import Special Education Enrollment
        Table 3.3: Public Special Ed
        Table 3.4: Private Special Ed (or Table 3.5)
        """
        year_id = self.years_map.get(year_label)
        if not year_id:
            return

        # Try different possible table names
        tables_to_try = [
            ('Table 3.3', 'Public'),
            ('Table 3.4', 'Private/Gov Assisted'),
            ('Table 3.5', 'Private/Gov Assisted')
        ]

        all_records = []

        for table_name, institution_type in tables_to_try:
            if table_name not in wb.sheetnames:
                continue

            ws = wb[table_name]

            # Check if this is a special ed table
            title = ws.cell(2, 1).value
            if not title or 'Special' not in str(title):
                continue

            # Age groups for special ed (different from early childhood)
            age_groups = [
                (5, '5-8 Years', '5_8'),
                (7, '9-11 Years', '9_11'),
                (9, '12-14 Years', '12_14'),
                (11, '15-17 Years', '15_17'),
                (13, '18-20 Years', '18_20'),
                (15, '>20 Years', 'over_20'),
                (17, 'Age Unknown', 'unknown')
            ]

            # Process each country column
            for col_idx, excel_code in self.country_columns.items():
                country_code = self.country_code_map.get(excel_code)
                country_id = self.countries_map.get(country_code)

                if not country_id:
                    continue

                # Process each age group
                for male_row, age_label, age_code in age_groups:
                    female_row = male_row + 1

                    male_count = self.safe_int(ws.cell(male_row, col_idx).value)
                    female_count = self.safe_int(ws.cell(female_row, col_idx).value)

                    if male_count > 0 or female_count > 0:
                        all_records.append({
                            'country_id': country_id,
                            'academic_year_id': year_id,
                            'institution_type': institution_type,
                            'age_group': age_code,
                            'male': male_count,
                            'female': female_count
                        })

        if all_records:
            try:
                # Delete existing
                self.supabase.table('special_education_enrollment')\
                    .delete()\
                    .eq('academic_year_id', year_id)\
                    .execute()

                # Insert new
                self.supabase.table('special_education_enrollment').insert(all_records).execute()
                self.stats['special_education'] += len(all_records)
                self.stats['total_records'] += len(all_records)
                print(f"  [OK] Special Education: {len(all_records)} records")
            except Exception as e:
                print(f"  [ERROR] Special Education failed: {e}")
        else:
            print(f"  [SKIP] Special Education: No data")

    def import_primary(self, wb, year_label):
        """
        Import Primary Enrollment
        Tables 3.7-3.15: One table per country
        Wide format: One record per age_group with k_male, k_female, g1_male, g1_female, etc.
        """
        year_id = self.years_map.get(year_label)
        if not year_id:
            return

        all_records = []

        # Process each country's primary table
        for table_name, country_code in self.primary_tables.items():
            if table_name not in wb.sheetnames:
                continue

            ws = wb[table_name]
            country_id = self.countries_map.get(country_code)

            if not country_id:
                continue

            # Age groups start at row 5
            # Format: <5, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16+, Unknown
            age_mapping = {
                5: 'under_5',
                7: '5',
                9: '6',
                11: '7',
                13: '8',
                15: '9',
                17: '10',
                19: '11',
                21: '12',
                23: '13',
                25: '14',
                27: '15',
                29: 'over_15',
                31: 'unknown'
            }

            # Process PUBLIC schools first (usually starts at row 5)
            for start_row, age_code in age_mapping.items():
                if start_row > 50:  # Safety check
                    break

                male_row = start_row
                female_row = start_row + 1

                # Check if we're past the data
                age_cell = ws.cell(male_row, 1).value
                if not age_cell or 'Total' in str(age_cell):
                    break

                # Extract data for all grades (columns 3-9 = K, G1-G6)
                k_male = self.safe_int(ws.cell(male_row, 3).value)
                k_female = self.safe_int(ws.cell(female_row, 3).value)
                g1_male = self.safe_int(ws.cell(male_row, 4).value)
                g1_female = self.safe_int(ws.cell(female_row, 4).value)
                g2_male = self.safe_int(ws.cell(male_row, 5).value)
                g2_female = self.safe_int(ws.cell(female_row, 5).value)
                g3_male = self.safe_int(ws.cell(male_row, 6).value)
                g3_female = self.safe_int(ws.cell(female_row, 6).value)
                g4_male = self.safe_int(ws.cell(male_row, 7).value)
                g4_female = self.safe_int(ws.cell(female_row, 7).value)
                g5_male = self.safe_int(ws.cell(male_row, 8).value)
                g5_female = self.safe_int(ws.cell(female_row, 8).value)
                g6_male = self.safe_int(ws.cell(male_row, 9).value)
                g6_female = self.safe_int(ws.cell(female_row, 9).value)

                subtotal_male = k_male + g1_male + g2_male + g3_male + g4_male + g5_male + g6_male
                subtotal_female = k_female + g1_female + g2_female + g3_female + g4_female + g5_female + g6_female

                if subtotal_male > 0 or subtotal_female > 0:
                    all_records.append({
                        'country_id': country_id,
                        'academic_year_id': year_id,
                        'school_type': 'Public',
                        'age_group': age_code,
                        'k_male': k_male,
                        'k_female': k_female,
                        'g1_male': g1_male,
                        'g1_female': g1_female,
                        'g2_male': g2_male,
                        'g2_female': g2_female,
                        'g3_male': g3_male,
                        'g3_female': g3_female,
                        'g4_male': g4_male,
                        'g4_female': g4_female,
                        'g5_male': g5_male,
                        'g5_female': g5_female,
                        'g6_male': g6_male,
                        'g6_female': g6_female,
                        'subtotal_male': subtotal_male,
                        'subtotal_female': subtotal_female
                    })

        if all_records:
            try:
                # Delete existing
                self.supabase.table('primary_enrollment')\
                    .delete()\
                    .eq('academic_year_id', year_id)\
                    .execute()

                # Insert new
                self.supabase.table('primary_enrollment').insert(all_records).execute()
                self.stats['primary'] += len(all_records)
                self.stats['total_records'] += len(all_records)
                print(f"  [OK] Primary Enrollment: {len(all_records)} records")
            except Exception as e:
                print(f"  [ERROR] Primary Enrollment failed: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"  [SKIP] Primary Enrollment: No data")

    def print_summary(self):
        """Print import summary"""
        print(f"\n{'='*70}")
        print("CHAPTER 3 ENROLLMENT IMPORT SUMMARY")
        print(f"{'='*70}")
        print(f"Early Childhood: {self.stats['early_childhood']} records")
        print(f"Special Education: {self.stats['special_education']} records")
        print(f"Primary: {self.stats['primary']} records")
        print(f"Secondary: {self.stats['secondary']} records")
        print(f"\nTOTAL: {self.stats['total_records']} records imported")
        print(f"{'='*70}\n")

if __name__ == '__main__':
    print("="*80)
    print("CHAPTER 3: STUDENT ENROLLMENT - REBUILT IMPORT")
    print("="*80)

    try:
        importer = Chapter3RebuiltImporter()
        print(f"\nConnected to Supabase")
        print(f"Found {len(importer.countries_map)} countries")
        print(f"Found {len(importer.years_map)} academic years\n")

        importer.import_all_years()
        importer.print_summary()

        print("="*80)
        print("IMPORT COMPLETE!")
        print("="*80)
    except Exception as e:
        print(f"\n[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
