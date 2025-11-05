"""
Import Chapter 1 (Institutions) Historical Data
Imports data from 2020-21, 2021-22, 2022-23 into the database
"""

import openpyxl
import os
from pathlib import Path

# Country code mapping (Excel row to database country_code)
COUNTRY_MAPPING = {
    'ANU': 'ANG',  # Anguilla
    'A&B': 'ATG',  # Antigua & Barbuda
    'DOM': 'DMA',  # Dominica
    'GRD': 'GRD',  # Grenada
    'MON': 'MSR',  # Montserrat
    'SKN': 'KNA',  # St. Kitts & Nevis
    'SLU': 'LCA',  # St. Lucia
    'SVG': 'VCT',  # St. Vincent & the Grenadines
    'VI': 'VGB'    # British Virgin Islands
}

# Excel row to country code mapping
ROW_TO_COUNTRY = {
    6: 'ANU',
    7: 'A&B',
    8: 'DOM',
    9: 'GRD',
    10: 'MON',
    11: 'SKN',
    12: 'SLU',
    13: 'SVG',
    14: 'VI'
}

class Chapter1Importer:
    def __init__(self, base_path):
        self.base_path = Path(base_path)
        self.data = []

    def import_all_years(self):
        """Import data from all three years"""
        print("=" * 70)
        print("CHAPTER 1 HISTORICAL DATA IMPORT")
        print("=" * 70)

        # Import each year
        years = [
            ('2020-21', '2020-2021', self.import_2020_21),
            ('2021-22', '2021-2022', self.import_2021_22),
            ('2022-23', '2022-2023', self.import_2022_23)
        ]

        for year_label, year_db, import_func in years:
            print(f"\nImporting {year_label}...")
            file_path = self.base_path / f"{year_label}.xlsx"

            if not file_path.exists():
                print(f"  [!] File not found: {file_path}")
                continue

            try:
                import_func(file_path, year_db)
                print(f"  [OK] {year_label} imported successfully")
            except Exception as e:
                print(f"  [ERROR] Error importing {year_label}: {e}")
                import traceback
                traceback.print_exc()

        return self.data

    def import_2020_21(self, file_path, year_label):
        """Import 2020-21 data (different sheet structure)"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 2020-21 uses different sheet names
        # Sheet "ECD" has daycare and preschool data
        if 'ECD' in wb.sheetnames:
            ws_ecd = wb['ECD']
            # Need to analyze the structure - may need adjustment
            print(f"    Found ECD sheet, analyzing structure...")

        # For now, mark as TODO since structure is different
        print(f"    [!] 2020-21 has different structure - needs manual mapping")
        print(f"    Sheet names: {wb.sheetnames}")

    def import_2021_22(self, file_path, year_label):
        """Import 2021-22 data"""
        self._import_standard_format(file_path, year_label)

    def import_2022_23(self, file_path, year_label):
        """Import 2022-23 data"""
        self._import_standard_format(file_path, year_label)

    def _import_standard_format(self, file_path, year_label):
        """Import data from standard format (2021-22, 2022-23)"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Process each country
        for row_num, country_code_excel in ROW_TO_COUNTRY.items():
            country_code_db = COUNTRY_MAPPING.get(country_code_excel)

            if not country_code_db:
                continue

            # Extract data from Table 1.1 (Early Childhood)
            ws_11 = wb['Table 1.1']
            daycare_public = ws_11[f'C{row_num}'].value or 0
            daycare_private_church = ws_11[f'D{row_num}'].value or 0
            daycare_private_non_affiliated = ws_11[f'E{row_num}'].value or 0
            preschool_public = ws_11[f'G{row_num}'].value or 0
            preschool_private_church = ws_11[f'H{row_num}'].value or 0
            preschool_private_non_affiliated = ws_11[f'I{row_num}'].value or 0

            # Extract data from Table 1.2 (Primary & Secondary)
            ws_12 = wb['Table 1.2']
            primary_public = ws_12[f'C{row_num}'].value or 0
            primary_private_church = ws_12[f'D{row_num}'].value or 0
            primary_private_non_affiliated = ws_12[f'E{row_num}'].value or 0
            secondary_public = ws_12[f'G{row_num}'].value or 0
            secondary_private_church = ws_12[f'H{row_num}'].value or 0
            secondary_private_non_affiliated = ws_12[f'I{row_num}'].value or 0

            # Special Education
            special_ed_public = ws_12[f'K{row_num}'].value or 0
            special_ed_private_church = ws_12[f'L{row_num}'].value or 0
            special_ed_private_non_affiliated = ws_12[f'M{row_num}'].value or 0

            # TVET (rows 17-25 in Table 1.2)
            row_num_tvet = row_num + 11  # TVET data starts 11 rows down
            tvet_public = ws_12[f'C{row_num_tvet}'].value or 0
            tvet_private_church = ws_12[f'D{row_num_tvet}'].value or 0
            tvet_private_non_affiliated = ws_12[f'E{row_num_tvet}'].value or 0

            # Extract data from Table 1.3 (Post-Secondary)
            ws_13 = wb['Table 1.3']
            # Post-secondary rows start at row 8
            row_num_ps = row_num + 0 if row_num >= 8 else row_num + 2
            post_secondary_public = ws_13[f'C{row_num_ps}'].value or 0
            post_secondary_private = ws_13[f'D{row_num_ps}'].value or 0

            # Helper function to safely convert to int
            def safe_int(value):
                if value is None:
                    return 0
                if isinstance(value, (int, float)):
                    return int(value)
                if isinstance(value, str):
                    # Remove special characters and try to convert
                    cleaned = value.strip().replace('�', '').replace('...', '').replace('-', '')
                    if cleaned == '' or cleaned == 'N/A':
                        return 0
                    try:
                        return int(float(cleaned))
                    except:
                        return 0
                return 0

            # Create data record
            record = {
                'country_code': country_code_db,
                'academic_year': year_label,
                'daycare_public': safe_int(daycare_public),
                'daycare_private_church': safe_int(daycare_private_church),
                'daycare_private_non_affiliated': safe_int(daycare_private_non_affiliated),
                'preschool_public': safe_int(preschool_public),
                'preschool_private_church': safe_int(preschool_private_church),
                'preschool_private_non_affiliated': safe_int(preschool_private_non_affiliated),
                'primary_public': safe_int(primary_public),
                'primary_private_church': safe_int(primary_private_church),
                'primary_private_non_affiliated': safe_int(primary_private_non_affiliated),
                'secondary_public': safe_int(secondary_public),
                'secondary_private_church': safe_int(secondary_private_church),
                'secondary_private_non_affiliated': safe_int(secondary_private_non_affiliated),
                'special_ed_public': safe_int(special_ed_public),
                'special_ed_private_church': safe_int(special_ed_private_church),
                'special_ed_private_non_affiliated': safe_int(special_ed_private_non_affiliated),
                'tvet_public': safe_int(tvet_public),
                'tvet_private_church': safe_int(tvet_private_church),
                'tvet_private_non_affiliated': safe_int(tvet_private_non_affiliated),
                'post_secondary_public': safe_int(post_secondary_public),
                'post_secondary_private': safe_int(post_secondary_private)
            }

            self.data.append(record)
            print(f"    [+] {country_code_db} ({year_label}): {sum([v for k,v in record.items() if k not in ['country_code', 'academic_year']])} total institutions")

    def generate_sql(self):
        """Generate SQL INSERT statements"""
        if not self.data:
            print("\n[!] No data to import")
            return

        print(f"\n{'=' * 70}")
        print(f"GENERATING SQL FOR {len(self.data)} RECORDS")
        print(f"{'=' * 70}\n")

        sql_statements = []

        # First, ensure academic years exist
        sql_statements.append("""
-- Ensure academic years exist
INSERT INTO academic_years (year_label, start_year, end_year, is_active)
VALUES
    ('2020-2021', 2020, 2021, false),
    ('2021-2022', 2021, 2022, false),
    ('2022-2023', 2022, 2023, false)
ON CONFLICT (year_label) DO NOTHING;
""")

        # Generate INSERT statements for each record
        for record in self.data:
            sql = f"""
-- {record['country_code']} - {record['academic_year']}
INSERT INTO institutions (
    country_id,
    academic_year_id,
    daycare_public, daycare_private_church, daycare_private_non_affiliated,
    preschool_public, preschool_private_church, preschool_private_non_affiliated,
    primary_public, primary_private_church, primary_private_non_affiliated,
    secondary_public, secondary_private_church, secondary_private_non_affiliated,
    special_ed_public, special_ed_private_church, special_ed_private_non_affiliated,
    tvet_public, tvet_private_church, tvet_private_non_affiliated,
    post_secondary_public, post_secondary_private
)
VALUES (
    (SELECT id FROM countries WHERE country_code = '{record['country_code']}'),
    (SELECT id FROM academic_years WHERE year_label = '{record['academic_year']}'),
    {record['daycare_public']}, {record['daycare_private_church']}, {record['daycare_private_non_affiliated']},
    {record['preschool_public']}, {record['preschool_private_church']}, {record['preschool_private_non_affiliated']},
    {record['primary_public']}, {record['primary_private_church']}, {record['primary_private_non_affiliated']},
    {record['secondary_public']}, {record['secondary_private_church']}, {record['secondary_private_non_affiliated']},
    {record['special_ed_public']}, {record['special_ed_private_church']}, {record['special_ed_private_non_affiliated']},
    {record['tvet_public']}, {record['tvet_private_church']}, {record['tvet_private_non_affiliated']},
    {record['post_secondary_public']}, {record['post_secondary_private']}
)
ON CONFLICT (country_id, academic_year_id)
DO UPDATE SET
    daycare_public = EXCLUDED.daycare_public,
    daycare_private_church = EXCLUDED.daycare_private_church,
    daycare_private_non_affiliated = EXCLUDED.daycare_private_non_affiliated,
    preschool_public = EXCLUDED.preschool_public,
    preschool_private_church = EXCLUDED.preschool_private_church,
    preschool_private_non_affiliated = EXCLUDED.preschool_private_non_affiliated,
    primary_public = EXCLUDED.primary_public,
    primary_private_church = EXCLUDED.primary_private_church,
    primary_private_non_affiliated = EXCLUDED.primary_private_non_affiliated,
    secondary_public = EXCLUDED.secondary_public,
    secondary_private_church = EXCLUDED.secondary_private_church,
    secondary_private_non_affiliated = EXCLUDED.secondary_private_non_affiliated,
    special_ed_public = EXCLUDED.special_ed_public,
    special_ed_private_church = EXCLUDED.special_ed_private_church,
    special_ed_private_non_affiliated = EXCLUDED.special_ed_private_non_affiliated,
    tvet_public = EXCLUDED.tvet_public,
    tvet_private_church = EXCLUDED.tvet_private_church,
    tvet_private_non_affiliated = EXCLUDED.tvet_private_non_affiliated,
    post_secondary_public = EXCLUDED.post_secondary_public,
    post_secondary_private = EXCLUDED.post_secondary_private,
    updated_at = NOW();
"""
            sql_statements.append(sql)

        return '\n'.join(sql_statements)


if __name__ == '__main__':
    # Path to Chapter 1 folder
    chapter1_path = r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chapter 1'

    # Create importer
    importer = Chapter1Importer(chapter1_path)

    # Import all years
    data = importer.import_all_years()

    # Generate SQL
    sql = importer.generate_sql()

    # Save to file
    output_file = r'C:\Users\Clendon\oecs-education-dashboard\import_chapter1_historical.sql'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(sql)

    print(f"\n{'=' * 70}")
    print(f"[SUCCESS] SQL file generated: {output_file}")
    print(f"{'=' * 70}")
    print(f"\nNext steps:")
    print(f"1. Review the SQL file")
    print(f"2. Run it in Supabase SQL Editor")
    print(f"3. Verify data imported correctly")
    print(f"4. Build trend dashboard!")
