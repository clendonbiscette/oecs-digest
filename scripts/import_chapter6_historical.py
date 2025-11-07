"""
Import Chapter 6 (Financial Data) Historical Data
Imports financial data from 2020-21, 2021-22, 2022-23 into the database

Table 6.1: National Budget and Education Expenditure by Country
"""
import openpyxl
from pathlib import Path

class Chapter6Importer:
    def __init__(self):
        self.base_path = Path(r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chp 6')
        self.data = []

        # Country code mapping
        self.country_map = {
            'ANG': 'ANG',
            'ANGUILLA': 'ANG',
            'A&B': 'ATG',
            'ATG': 'ATG',
            'ANTIGUA': 'ATG',
            'DOM': 'DMA',
            'DMA': 'DMA',
            'DOMINICA': 'DMA',
            'GREN': 'GRD',
            'GRD': 'GRD',
            'GRENADA': 'GRD',
            'MON': 'MSR',
            'MSR': 'MSR',
            'MONTSERRAT': 'MSR',
            'SKN': 'KNA',
            'KNA': 'KNA',
            'ST. KITTS': 'KNA',
            'SLU': 'LCA',
            'LCA': 'LCA',
            'ST. LUCIA': 'LCA',
            'SVG': 'VCT',
            'VCT': 'VCT',
            'ST. VINCENT': 'VCT',
            'BVI': 'VGB',
            'VGB': 'VGB'
        }

    def safe_float(self, value):
        """Safely convert value to float"""
        if value is None or value == '':
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Remove special characters
            cleaned = value.strip().replace('�', '').replace(',', '').replace('...', '')
            if cleaned == '' or cleaned == 'N/A' or cleaned == '#DIV/0!':
                return None
            try:
                return float(cleaned)
            except:
                return None
        return None

    def import_all_years(self):
        """Import all available years"""
        years = {
            '2021-2022': ('2021-22', self.import_2021_22),
            '2022-2023': ('2022-23', self.import_2022_23),
            '2020-2021': ('2020-21', self.import_2020_21)
        }

        for year_db, (year_label, import_func) in years.items():
            file_path = self.base_path / f"{year_label}.xlsx"

            if not file_path.exists():
                print(f"  [!] File not found: {file_path}")
                continue

            try:
                import_func(file_path, year_db)
                print(f"  [OK] {year_label} imported successfully")
            except Exception as e:
                print(f"  [ERROR] Error importing {year_label}: {e}")
                import traceback
                traceback.print_exc()

    def import_2022_23(self, file_path, year_label):
        """Import 2022-23 data"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['Table 6.1']

        # Country columns: B=DOM, C=GREN, D=SKN, E=SLU, F=SVG
        countries = {
            'B': 'DMA',
            'C': 'GRD',
            'D': 'KNA',
            'E': 'LCA',
            'F': 'VCT'
        }

        for col, country_code in countries.items():
            col_idx = ord(col) - ord('A') + 1

            record = {
                'country_code': country_code,
                'academic_year': year_label,
                'national_budget_total': self.safe_float(ws.cell(2, col_idx).value),
                'national_budget_recurrent': self.safe_float(ws.cell(3, col_idx).value),
                'national_budget_capital': self.safe_float(ws.cell(4, col_idx).value),
                'gdp': self.safe_float(ws.cell(5, col_idx).value),
                'education_budget_total': self.safe_float(ws.cell(6, col_idx).value),
                'education_budget_recurrent': self.safe_float(ws.cell(7, col_idx).value),
                'education_budget_capital': self.safe_float(ws.cell(8, col_idx).value),
                'education_pct_national_budget': self.safe_float(ws.cell(9, col_idx).value),
                'education_pct_gdp': self.safe_float(ws.cell(10, col_idx).value),
                'allocation_early_childhood': self.safe_float(ws.cell(12, col_idx).value),
                'allocation_primary': self.safe_float(ws.cell(13, col_idx).value),
                'allocation_secondary': self.safe_float(ws.cell(14, col_idx).value),
                'allocation_special_ed': self.safe_float(ws.cell(15, col_idx).value),
                'allocation_post_secondary': self.safe_float(ws.cell(16, col_idx).value),
                'allocation_tertiary': self.safe_float(ws.cell(17, col_idx).value),
                'allocation_other': self.safe_float(ws.cell(18, col_idx).value)
            }

            self.data.append(record)
            budget_val = record['national_budget_total']
            budget_str = f"${budget_val:,.0f}" if budget_val else "$0"
            print(f"    [+] {country_code} ({year_label}): Budget={budget_str}")

    def import_2021_22(self, file_path, year_label):
        """Import 2021-22 data"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['Table 6.1']

        # Read header row to find country columns
        header_row = [cell.value for cell in ws[1]]

        # Map columns to countries
        country_columns = {}
        for idx, header in enumerate(header_row):
            if header and str(header).strip().upper() in self.country_map:
                country_code = self.country_map[str(header).strip().upper()]
                country_columns[idx] = country_code

        print(f"    Found countries: {list(country_columns.values())}")

        for col_idx, country_code in country_columns.items():
            record = {
                'country_code': country_code,
                'academic_year': year_label,
                'national_budget_total': self.safe_float(ws.cell(2, col_idx + 1).value),
                'national_budget_recurrent': self.safe_float(ws.cell(3, col_idx + 1).value),
                'national_budget_capital': self.safe_float(ws.cell(4, col_idx + 1).value),
                'gdp': self.safe_float(ws.cell(5, col_idx + 1).value),
                'education_budget_total': self.safe_float(ws.cell(6, col_idx + 1).value),
                'education_budget_recurrent': self.safe_float(ws.cell(7, col_idx + 1).value),
                'education_budget_capital': self.safe_float(ws.cell(8, col_idx + 1).value),
                'education_pct_national_budget': self.safe_float(ws.cell(9, col_idx + 1).value),
                'education_pct_gdp': self.safe_float(ws.cell(10, col_idx + 1).value),
                'allocation_early_childhood': self.safe_float(ws.cell(12, col_idx + 1).value),
                'allocation_primary': self.safe_float(ws.cell(13, col_idx + 1).value),
                'allocation_secondary': self.safe_float(ws.cell(14, col_idx + 1).value),
                'allocation_special_ed': self.safe_float(ws.cell(15, col_idx + 1).value),
                'allocation_post_secondary': self.safe_float(ws.cell(16, col_idx + 1).value),
                'allocation_tertiary': self.safe_float(ws.cell(17, col_idx + 1).value),
                'allocation_other': self.safe_float(ws.cell(18, col_idx + 1).value)
            }

            self.data.append(record)
            if record['national_budget_total']:
                print(f"    [+] {country_code} ({year_label}): Budget=${record['national_budget_total']:,.0f}")

    def import_2020_21(self, file_path, year_label):
        """Import 2020-21 data - different structure"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['Table 6.1']

        # Row 4 has headers: Description, A&B, DOM, GREN, MON, SKN, SLU, SVG, VI, OECS
        # Rows 5-20 have the data

        countries = {
            2: ('A&B', 'ATG'),      # Column B
            3: ('DOM', 'DMA'),       # Column C
            4: ('GREN', 'GRD'),      # Column D
            5: ('MON', 'MSR'),       # Column E
            6: ('SKN', 'KNA'),       # Column F
            7: ('SLU', 'LCA'),       # Column G
            8: ('SVG', 'VCT'),       # Column H
            9: ('VI', 'VGB')         # Column I
        }

        for col_idx, (header, country_code) in countries.items():
            record = {
                'country_code': country_code,
                'academic_year': year_label,
                'national_budget_total': self.safe_float(ws.cell(5, col_idx).value),
                'national_budget_recurrent': self.safe_float(ws.cell(6, col_idx).value),
                'national_budget_capital': self.safe_float(ws.cell(7, col_idx).value),
                'gdp': self.safe_float(ws.cell(8, col_idx).value),
                'education_budget_total': self.safe_float(ws.cell(9, col_idx).value),
                'education_budget_recurrent': self.safe_float(ws.cell(10, col_idx).value),
                'education_budget_capital': self.safe_float(ws.cell(11, col_idx).value),
                'education_pct_national_budget': self.safe_float(ws.cell(12, col_idx).value),
                'education_pct_gdp': self.safe_float(ws.cell(13, col_idx).value),
                'allocation_early_childhood': self.safe_float(ws.cell(16, col_idx).value),
                'allocation_primary': self.safe_float(ws.cell(17, col_idx).value),
                'allocation_secondary': self.safe_float(ws.cell(18, col_idx).value),
                'allocation_special_ed': self.safe_float(ws.cell(19, col_idx).value),
                'allocation_post_secondary': self.safe_float(ws.cell(20, col_idx).value),
                'allocation_tertiary': self.safe_float(ws.cell(21, col_idx).value) if ws.max_row >= 21 else None,
                'allocation_other': self.safe_float(ws.cell(22, col_idx).value) if ws.max_row >= 22 else None
            }

            self.data.append(record)
            if record['national_budget_total']:
                print(f"    [+] {country_code} ({year_label}): Budget=${record['national_budget_total']:,.0f}")

    def generate_sql(self):
        """Generate SQL INSERT statements"""
        if not self.data:
            print("\n[!] No data to import")
            return

        print(f"\n{'=' * 70}")
        print(f"GENERATING SQL FOR {len(self.data)} RECORDS")
        print(f"{'=' * 70}\n")

        sql = "-- Chapter 6: Financial Data Import\n"
        sql += "-- Generated from historical Excel files\n\n"

        # Create table if not exists
        sql += """
-- Create financial_data table
CREATE TABLE IF NOT EXISTS financial_data (
    id SERIAL PRIMARY KEY,
    country_id INTEGER REFERENCES countries(id),
    academic_year_id INTEGER REFERENCES academic_years(id),
    national_budget_total DECIMAL(15,2),
    national_budget_recurrent DECIMAL(15,2),
    national_budget_capital DECIMAL(15,2),
    gdp DECIMAL(15,2),
    education_budget_total DECIMAL(15,2),
    education_budget_recurrent DECIMAL(15,2),
    education_budget_capital DECIMAL(15,2),
    education_pct_national_budget DECIMAL(5,2),
    education_pct_gdp DECIMAL(5,2),
    allocation_early_childhood DECIMAL(15,2),
    allocation_primary DECIMAL(15,2),
    allocation_secondary DECIMAL(15,2),
    allocation_special_ed DECIMAL(15,2),
    allocation_post_secondary DECIMAL(15,2),
    allocation_tertiary DECIMAL(15,2),
    allocation_other DECIMAL(15,2),
    created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
    updated_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
    UNIQUE(country_id, academic_year_id)
);

-- Create indexes
CREATE INDEX IF NOT EXISTS idx_financial_country ON financial_data(country_id);
CREATE INDEX IF NOT EXISTS idx_financial_year ON financial_data(academic_year_id);

-- Enable RLS
ALTER TABLE financial_data ENABLE ROW LEVEL SECURITY;

-- RLS Policies
DROP POLICY IF EXISTS "Users can view financial data" ON financial_data;
CREATE POLICY "Users can view financial data" ON financial_data
    FOR SELECT TO authenticated USING (true);

GRANT SELECT ON financial_data TO authenticated;

"""

        # Insert data
        for record in self.data:
            country_code = record['country_code']
            year = record['academic_year']

            # Build values list
            values = []
            for key in [
                'national_budget_total', 'national_budget_recurrent', 'national_budget_capital',
                'gdp', 'education_budget_total', 'education_budget_recurrent',
                'education_budget_capital', 'education_pct_national_budget', 'education_pct_gdp',
                'allocation_early_childhood', 'allocation_primary', 'allocation_secondary',
                'allocation_special_ed', 'allocation_post_secondary', 'allocation_tertiary',
                'allocation_other'
            ]:
                val = record[key]
                if val is None:
                    values.append('NULL')
                else:
                    values.append(f"{val}")

            sql += f"""
-- {country_code} - {year}
INSERT INTO financial_data (
    country_id,
    academic_year_id,
    national_budget_total, national_budget_recurrent, national_budget_capital,
    gdp, education_budget_total, education_budget_recurrent, education_budget_capital,
    education_pct_national_budget, education_pct_gdp,
    allocation_early_childhood, allocation_primary, allocation_secondary,
    allocation_special_ed, allocation_post_secondary, allocation_tertiary, allocation_other
)
VALUES (
    (SELECT id FROM countries WHERE country_code = '{country_code}'),
    (SELECT id FROM academic_years WHERE year_label = '{year}'),
    {', '.join(values)}
)
ON CONFLICT (country_id, academic_year_id)
DO UPDATE SET
    national_budget_total = EXCLUDED.national_budget_total,
    national_budget_recurrent = EXCLUDED.national_budget_recurrent,
    national_budget_capital = EXCLUDED.national_budget_capital,
    gdp = EXCLUDED.gdp,
    education_budget_total = EXCLUDED.education_budget_total,
    education_budget_recurrent = EXCLUDED.education_budget_recurrent,
    education_budget_capital = EXCLUDED.education_budget_capital,
    education_pct_national_budget = EXCLUDED.education_pct_national_budget,
    education_pct_gdp = EXCLUDED.education_pct_gdp,
    allocation_early_childhood = EXCLUDED.allocation_early_childhood,
    allocation_primary = EXCLUDED.allocation_primary,
    allocation_secondary = EXCLUDED.allocation_secondary,
    allocation_special_ed = EXCLUDED.allocation_special_ed,
    allocation_post_secondary = EXCLUDED.allocation_post_secondary,
    allocation_tertiary = EXCLUDED.allocation_tertiary,
    allocation_other = EXCLUDED.allocation_other;
"""

        # Write to file
        output_file = 'import_chapter6_historical.sql'
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(sql)

        print(f"[SUCCESS] SQL file generated: {output_file}")
        print(f"Total records: {len(self.data)}")

if __name__ == '__main__':
    print("=" * 70)
    print("CHAPTER 6 HISTORICAL DATA IMPORT - FINANCIAL DATA")
    print("=" * 70)
    print()

    importer = Chapter6Importer()
    importer.import_all_years()
    importer.generate_sql()

    print(f"\n{'=' * 70}")
    print("IMPORT COMPLETE")
    print(f"{'=' * 70}")
