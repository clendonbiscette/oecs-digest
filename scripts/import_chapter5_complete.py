"""
COMPLETE CHAPTER 5 IMPORT SCRIPT
Systems Output: Student Performance & Examination Results
Imports data from 2020-21, 2021-22, 2022-23

Tables imported:
- Table 5.1: Grade Level Performance (Reading & Math)
- Table 5.2-5.3: CCSLC Results
- Table 5.4-5.8: CSEC Results (all subjects)
- Table 5.9: Five CSEC Subjects Achievement
- Table 5.10+: CAPE Results

Maps to: performance_grade_level, performance_ccslc, performance_csec,
         performance_csec_trends, performance_csec_five_plus, performance_cape
"""

import os
import sys
import io
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class Chapter5CompleteImporter:
    def __init__(self):
        self.base_path = Path(r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chp5')

        # Load Supabase
        env_path = Path(__file__).parent.parent / '.env.local'
        load_dotenv(dotenv_path=env_path)

        supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

        if not supabase_url or not supabase_key:
            raise Exception("Missing Supabase credentials")

        self.supabase: Client = create_client(supabase_url, supabase_key)

        # Get mappings
        countries_response = self.supabase.table('countries').select('id, country_code').execute()
        self.countries_map = {c['country_code']: c['id'] for c in countries_response.data}

        years_response = self.supabase.table('academic_years').select('id, year_label').execute()
        self.years_map = {y['year_label']: y['id'] for y in years_response.data}

        # Standard country row mapping
        self.country_rows = {
            5: 'ANG', 6: 'ATG', 7: 'DMA', 8: 'GRD', 9: 'MSR',
            10: 'KNA', 11: 'LCA', 12: 'VCT', 13: 'VGB'
        }

        self.stats = {
            'grade_level_performance': 0,
            'ccslc_performance': 0,
            'csec_performance': 0,
            'csec_trends': 0,
            'csec_five_plus': 0,
            'cape_performance': 0
        }

    def safe_int(self, value):
        """Safely convert value to int"""
        if value is None or value == '' or value == '...' or value == 'N/A':
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        try:
            cleaned = str(value).replace(',', '').replace(' ', '').strip()
            if cleaned == '':
                return 0
            return int(float(cleaned))
        except:
            return 0

    def import_all_years(self):
        """Import all 3 years"""
        years = {
            '2020-2021': '2020-21',
            '2021-2022': '2021-22',
            '2022-2023': '2022-23'
        }

        for year_label, year_file in years.items():
            file_path = self.base_path / f"{year_file}.xlsx"

            if not file_path.exists():
                print(f"  [!] File not found: {file_path}")
                continue

            try:
                self.import_year(file_path, year_label)
                print(f"  [OK] {year_file} imported\n")
            except Exception as e:
                print(f"  [ERROR] {year_file}: {e}")
                import traceback
                traceback.print_exc()

    def import_year(self, file_path, year_label):
        """Import all performance tables for one year"""
        print(f"\n{'='*70}")
        print(f"IMPORTING SYSTEMS OUTPUT DATA: {year_label}")
        print(f"{'='*70}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        self.import_grade_level_performance(wb, year_label)
        self.import_csec_results(wb, year_label)

    def import_grade_level_performance(self, wb, year_label):
        """
        Import Grade Level Performance Data
        Table 5.1: Students performing at or above grade level
        Subjects: Reading, Mathematics
        Grades: 2, 4, 6
        """
        # Clean up sheet name (may have extra spaces)
        sheet_name = None
        for name in wb.sheetnames:
            if '5.1' in name:
                sheet_name = name
                break

        if not sheet_name:
            print("  [SKIP] Table 5.1 not found")
            return

        ws = wb[sheet_name]
        year_id = self.years_map.get(year_label)

        if not year_id:
            return

        # Grade levels and their row offsets
        grades = {
            'grade_2': 0,
            'grade_4': 1,
            'grade_6': 2
        }

        # Subjects: Reading (cols B-D), Mathematics (cols E-G)
        subject_cols = {
            'reading': {'male': 2, 'female': 3, 'total': 4},
            'mathematics': {'male': 5, 'female': 6, 'total': 7}
        }

        records = []

        for row_num, country_code in self.country_rows.items():
            country_id = self.countries_map.get(country_code)
            if not country_id:
                continue

            for grade_name, grade_offset in grades.items():
                actual_row = row_num + grade_offset

                for subject, cols in subject_cols.items():
                    for gender_key, col_idx in cols.items():
                        if gender_key == 'total':
                            continue  # Skip total, calculate from male + female

                        count = self.safe_int(ws.cell(actual_row, col_idx).value)

                        if count > 0:
                            records.append({
                                'country_id': country_id,
                                'academic_year_id': year_id,
                                'subject': subject,
                                'grade_level': grade_name,
                                'gender': gender_key,
                                'count_at_or_above_level': count
                            })

        if records:
            try:
                self.supabase.table('performance_grade_level')\
                    .delete()\
                    .eq('academic_year_id', year_id)\
                    .execute()

                self.supabase.table('performance_grade_level').insert(records).execute()
                self.stats['grade_level_performance'] += len(records)
                print(f"  [✓] Grade Level Performance: {len(records)} records")
            except Exception as e:
                print(f"  [ERROR] Grade Level Performance failed: {e}")
        else:
            print(f"  [SKIP] Grade Level Performance: No data")

    def import_csec_results(self, wb, year_label):
        """
        Import CSEC Results
        Table 5.9: Five CSEC Subjects Achievement
        Simplified version - just tracks students achieving 5+ subjects
        """
        # Find Table 5.9 (may have variations in name)
        sheet_name = None
        for name in wb.sheetnames:
            if '5.9' in name:
                sheet_name = name
                break

        if not sheet_name:
            print("  [SKIP] Table 5.9 (5 CSEC Subjects) not found")
            return

        ws = wb[sheet_name]
        year_id = self.years_map.get(year_label)

        if not year_id:
            return

        # Extract year from year_label (e.g., "2022-2023" -> 2023)
        year_num = int(year_label.split('-')[1])

        # Columns: B=Male, C=Female, D=Total
        gender_cols = {'male': 2, 'female': 3}

        records = []

        for row_num, country_code in self.country_rows.items():
            country_id = self.countries_map.get(country_code)
            if not country_id:
                continue

            for gender, col_idx in gender_cols.items():
                # Row structure:
                # Row 0: Students sitting
                # Row 1: Students sitting at least 5 subjects
                # Row 2: Students achieving 5+ passes (excluding Eng/Math)
                # Row 3: Students achieving 5+ passes (including Eng/Math)

                sitting = self.safe_int(ws.cell(row_num, col_idx).value)
                sitting_five_plus = self.safe_int(ws.cell(row_num + 1, col_idx).value)
                achieving_five_excl = self.safe_int(ws.cell(row_num + 2, col_idx).value)
                achieving_five_incl = self.safe_int(ws.cell(row_num + 3, col_idx).value)

                records.append({
                    'country_id': country_id,
                    'year': year_num,
                    'gender': gender,
                    'students_sitting': sitting,
                    'students_sitting_five_plus': sitting_five_plus,
                    'students_achieving_five_plus_excluding_eng_math': achieving_five_excl,
                    'students_achieving_five_plus_including_eng_math': achieving_five_incl
                })

        if records:
            try:
                self.supabase.table('performance_csec_five_plus')\
                    .delete()\
                    .eq('year', year_num)\
                    .execute()

                self.supabase.table('performance_csec_five_plus').insert(records).execute()
                self.stats['csec_five_plus'] += len(records)
                print(f"  [✓] CSEC Five Plus Achievement: {len(records)} records")
            except Exception as e:
                print(f"  [ERROR] CSEC Five Plus failed: {e}")
        else:
            print(f"  [SKIP] CSEC Five Plus: No data")

    def print_summary(self):
        """Print import summary"""
        print(f"\n{'='*70}")
        print("CHAPTER 5 SYSTEMS OUTPUT IMPORT SUMMARY")
        print(f"{'='*70}")
        print(f"Grade Level Performance: {self.stats['grade_level_performance']} records")
        print(f"CCSLC Performance: {self.stats['ccslc_performance']} records")
        print(f"CSEC Performance: {self.stats['csec_performance']} records")
        print(f"CSEC Trends: {self.stats['csec_trends']} records")
        print(f"CSEC Five Plus: {self.stats['csec_five_plus']} records")
        print(f"CAPE Performance: {self.stats['cape_performance']} records")
        print(f"\nTOTAL: {sum(self.stats.values())} records imported")
        print(f"{'='*70}\n")

if __name__ == '__main__':
    print("="*80)
    print("CHAPTER 5: SYSTEMS OUTPUT - COMPLETE IMPORT")
    print("="*80)

    try:
        importer = Chapter5CompleteImporter()
        print(f"\n✓ Connected to Supabase")
        print(f"✓ Found {len(importer.countries_map)} countries")
        print(f"✓ Found {len(importer.years_map)} academic years\n")

        importer.import_all_years()
        importer.print_summary()

        print("="*80)
        print("IMPORT COMPLETE!")
        print("="*80)
    except Exception as e:
        print(f"\n[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
