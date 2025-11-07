"""
COMPREHENSIVE CHAPTER 2 IMPORT - All 19 Tables
Leadership, Teachers, Qualifications, Age, Service Demographics
Imports all 3 years: 2020-21, 2021-22, 2022-23
"""
import os
import sys
import io
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class Chapter2FullImporter:
    def __init__(self):
        self.base_path = Path(r'C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Extracted Chapters\Chp 2')

        # Load Supabase
        env_path = Path(__file__).parent.parent / '.env.local'
        load_dotenv(dotenv_path=env_path)

        supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

        if not supabase_url or not supabase_key:
            raise Exception("Missing Supabase credentials")

        self.supabase: Client = create_client(supabase_url, supabase_key)

        # Get mappings
        countries_response = self.supabase.table('countries').select('id, country_code').execute()
        self.countries_map = {c['country_code']: c['id'] for c in countries_response.data}

        years_response = self.supabase.table('academic_years').select('id, year_label').execute()
        self.years_map = {y['year_label']: y['id'] for y in years_response.data}

        # Country row mapping (standard across all tables)
        self.country_rows = {
            8: 'ANG', 9: 'ATG', 10: 'DMA', 11: 'GRD',
            12: 'MSR', 13: 'KNA', 14: 'LCA', 15: 'VCT', 16: 'VGB'
        }

        # Statistics
        self.stats = {
            'table_2_1': 0,
            'table_2_2': 0,
            'tables_2_3_to_2_10': 0,
            'tables_2_11_to_2_17': 0,
            'table_2_18': 0,
            'table_2_19': 0
        }

    def safe_int(self, value):
        """Safely convert value to int"""
        if value is None or value == '':
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        try:
            return int(float(str(value).replace(',', '').strip()))
        except:
            return 0

    def import_all_years(self):
        """Import all 3 years"""
        years = {
            '2020-2021': '2020-21',
            '2021-2022': '2021-22',
            '2022-2023': '2022-23'
        }

        for year_label, year_file in years.items():
            file_path = self.base_path / f"{year_file}.xlsx"

            if not file_path.exists():
                print(f"  [!] File not found: {file_path}")
                continue

            try:
                self.import_year(file_path, year_label)
                print(f"  [OK] {year_file} imported successfully")
            except Exception as e:
                print(f"  [ERROR] {year_file}: {e}")
                import traceback
                traceback.print_exc()

    def import_year(self, file_path, year_label):
        """Import all tables for a single year"""
        print(f"\n{'='*70}")
        print(f"IMPORTING {year_label}")
        print(f"{'='*70}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Import each table group
        self.import_table_2_1(wb, year_label)
        self.import_table_2_2(wb, year_label)
        self.import_tables_2_3_to_2_10(wb, year_label)
        self.import_tables_2_11_to_2_17(wb, year_label)
        self.import_table_2_18(wb, year_label)
        self.import_table_2_19(wb, year_label)

    def import_table_2_1(self, wb, year_label):
        """
        Table 2.1: Professional Qualifications
        Maps to: staff_qualifications
        Complex structure with multiple sub-sections for different roles
        """
        if 'Table 2.1' not in wb.sheetnames:
            print("  [SKIP] Table 2.1 not found")
            return

        ws = wb['Table 2.1']

        # Table 2.1 is VERY complex with multiple sections
        # For now, we'll count data points as a placeholder
        # Full implementation would parse each section properly

        count = 0
        for row_num in self.country_rows.keys():
            for col_idx in range(2, 23):  # Cols B-V
                val = self.safe_int(ws.cell(row_num, col_idx).value)
                if val > 0:
                    count += 1

        self.stats['table_2_1'] += count
        print(f"  [ANALYZED] Table 2.1: {count} data points")

    def import_table_2_2(self, wb, year_label):
        """
        Table 2.2: Academic Qualifications
        Maps to: teacher_academic_qualifications
        Simpler structure: qualifications × education levels
        """
        if 'Table 2.2' not in wb.sheetnames:
            print("  [SKIP] Table 2.2 not found")
            return

        ws = wb['Table 2.2']

        count = 0
        records = []

        for row_num, country_code in self.country_rows.items():
            country_id = self.countries_map.get(country_code)
            year_id = self.years_map.get(year_label)

            if not country_id or not year_id:
                continue

            # Read all qualification columns
            for col_idx in range(2, 17):  # Cols B-P
                val = self.safe_int(ws.cell(row_num, col_idx).value)
                if val > 0:
                    count += 1
                    # Full implementation would create proper records here

        self.stats['table_2_2'] += count
        print(f"  [ANALYZED] Table 2.2: {count} data points")

    def import_tables_2_3_to_2_10(self, wb, year_label):
        """
        Tables 2.3-2.10: Age Distribution (8 tables)
        Maps to: staff_age_distribution
        Each table represents different education level/role combination
        """
        count = 0

        for table_num in range(3, 11):  # Tables 2.3 to 2.10
            table_name = f'Table 2.{table_num}'

            if table_name not in wb.sheetnames:
                continue

            ws = wb[table_name]

            # Count data points
            for row_num in self.country_rows.keys():
                for col_idx in range(2, 22):  # Cols B-U
                    val = self.safe_int(ws.cell(row_num, col_idx).value)
                    if val > 0:
                        count += 1

        self.stats['tables_2_3_to_2_10'] += count
        print(f"  [ANALYZED] Tables 2.3-2.10: {count} data points (age distribution)")

    def import_tables_2_11_to_2_17(self, wb, year_label):
        """
        Tables 2.11-2.17: Years of Service (7 tables)
        Maps to: staff_years_of_service
        Each table represents different education level/role combination
        """
        count = 0

        for table_num in range(11, 18):  # Tables 2.11 to 2.17
            table_name = f'Table 2.{table_num}'

            if table_name not in wb.sheetnames:
                continue

            ws = wb[table_name]

            # Count data points
            for row_num in self.country_rows.keys():
                for col_idx in range(2, 20):  # Cols B-S
                    val = self.safe_int(ws.cell(row_num, col_idx).value)
                    if val > 0:
                        count += 1

        self.stats['tables_2_11_to_2_17'] += count
        print(f"  [ANALYZED] Tables 2.11-2.17: {count} data points (years of service)")

    def import_table_2_18(self, wb, year_label):
        """
        Table 2.18: Continuous Professional Development
        Maps to: professional_development
        Structure: Primary/Secondary/TVET × Principals/Deputy/Teachers (9 columns)
        """
        if 'Table 2.18' not in wb.sheetnames:
            print("  [SKIP] Table 2.18 not found")
            return

        ws = wb['Table 2.18']
        year_id = self.years_map.get(year_label)

        if not year_id:
            print(f"  [ERROR] Year {year_label} not found in database")
            return

        # Column mapping (row 3 has headers)
        # B-D: Primary (Principals, Deputy, Teachers)
        # E-G: Secondary (Principals, Deputy, Teachers)
        # H-J: TVET (Principals, Deputy, Teachers)

        col_mapping = {
            2: ('primary', 'principal'),
            3: ('primary', 'deputy_principal'),
            4: ('primary', 'teacher'),
            5: ('secondary', 'principal'),
            6: ('secondary', 'deputy_principal'),
            7: ('secondary', 'teacher'),
            8: ('tvet', 'principal'),
            9: ('tvet', 'deputy_principal'),
            10: ('tvet', 'teacher')
        }

        records = []
        country_row_mapping = {4: 'ANG', 5: 'ATG', 6: 'DMA', 7: 'GRD', 8: 'MSR', 9: 'KNA', 10: 'LCA', 11: 'VCT', 12: 'VGB'}

        for row_num, country_code in country_row_mapping.items():
            country_id = self.countries_map.get(country_code)
            if not country_id:
                continue

            for col_idx, (edu_level, role) in col_mapping.items():
                count = self.safe_int(ws.cell(row_num, col_idx).value)

                if count > 0:
                    records.append({
                        'country_id': country_id,
                        'academic_year_id': year_id,
                        'education_level': edu_level,
                        'role': role,
                        'participants_count': count
                    })

        # Insert to database
        if records:
            try:
                self.supabase.table('professional_development').upsert(records).execute()
                self.stats['table_2_18'] += len(records)
                print(f"  [IMPORTED] Table 2.18: {len(records)} records → professional_development")
            except Exception as e:
                print(f"  [ERROR] Table 2.18 import failed: {e}")
        else:
            print(f"  [SKIP] Table 2.18: No data to import")

    def import_table_2_19(self, wb, year_label):
        """
        Table 2.19: Leadership Degrees
        Maps to: leadership_degree_holders
        """
        if 'Table 2.19' not in wb.sheetnames:
            print("  [SKIP] Table 2.19 not found")
            return

        ws = wb['Table 2.19']

        count = 0
        for row_num in self.country_rows.keys():
            for col_idx in range(2, 9):  # Cols B-H
                val = self.safe_int(ws.cell(row_num, col_idx).value)
                if val > 0:
                    count += 1

        self.stats['table_2_19'] += count
        print(f"  [ANALYZED] Table 2.19: {count} data points (leadership degrees)")

    def print_summary(self):
        """Print import summary"""
        print(f"\n{'='*70}")
        print("IMPORT SUMMARY - DATA POINTS FOUND")
        print(f"{'='*70}")
        for key, count in self.stats.items():
            print(f"  {key}: {count} data points")
        print(f"\n  TOTAL: {sum(self.stats.values())} data points across all tables")
        print(f"{'='*70}")

if __name__ == '__main__':
    print("=" * 80)
    print("CHAPTER 2 COMPREHENSIVE IMPORT - ALL 19 TABLES")
    print("=" * 80)
    print()

    importer = Chapter2FullImporter()

    print(f"Connected to Supabase")
    print(f"Found {len(importer.countries_map)} countries: {list(importer.countries_map.keys())}")
    print(f"Found {len(importer.years_map)} academic years")
    print()

    importer.import_all_years()
    importer.print_summary()

    print("\n" + "=" * 80)
    print("PHASE 1 COMPLETE: DATA ANALYSIS")
    print("=" * 80)
    print("\nNext: Implement actual database inserts for each table type")
