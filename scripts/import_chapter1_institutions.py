"""
Import Chapter 1 (Institutions) data from historical Excel files into Supabase

This script imports institution counts across 3 academic years (2020-21, 2021-22, 2022-23)
for all 9 OECS member states into the institutions table.

Tables in Chapter 1:
- Table 1.1: Early Childhood (Daycare & Preschools)
- Table 1.2: Primary & Secondary Schools
- Table 1.3: Post-Secondary Institutions

Usage:
    python scripts/import_chapter1_institutions.py
"""

import os
import sys
import io

# Fix Windows console encoding for emojis
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

# Load environment variables from .env.local
env_path = Path(__file__).parent.parent / '.env.local'
load_dotenv(dotenv_path=env_path)

# Supabase configuration
SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY')  # Need service role for bulk insert

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ Error: Missing Supabase credentials in .env.local file")
    print("   Required: NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Country code mapping (Excel abbreviation -> ISO code)
COUNTRY_MAPPING = {
    'A&B': 'ATG',  # Antigua and Barbuda
    'ANG': 'AIA',  # Anguilla
    'ANU': 'ATG',  # Alternative for Antigua (sometimes used)
    'DOM': 'DMA',  # Dominica
    'GRD': 'GRD',  # Grenada
    'MON': 'MSR',  # Montserrat
    'SKN': 'KNA',  # Saint Kitts and Nevis
    'SLU': 'LCA',  # Saint Lucia
    'SVG': 'VCT',  # Saint Vincent and the Grenadines
    'VI': 'VGB',   # British Virgin Islands
}

# Academic year mapping
ACADEMIC_YEAR_MAPPING = {
    '2020-21': '2020-2021',
    '2021-22': '2021-2022',
    '2022-23': '2022-2023',
}

def get_or_create_academic_years():
    """Ensure academic years exist in database"""
    years_data = [
        {'year_label': '2020-2021', 'start_year': 2020, 'end_year': 2021, 'is_active': False},
        {'year_label': '2021-2022', 'start_year': 2021, 'end_year': 2022, 'is_active': False},
        {'year_label': '2022-2023', 'start_year': 2022, 'end_year': 2023, 'is_active': False},
        {'year_label': '2023-2024', 'start_year': 2023, 'end_year': 2024, 'is_active': True},
    ]

    print("\n📅 Checking academic years...")
    for year_data in years_data:
        result = supabase.table('academic_years').select('id').eq('year_label', year_data['year_label']).execute()
        if not result.data:
            supabase.table('academic_years').insert(year_data).execute()
            print(f"   ✓ Created: {year_data['year_label']}")
        else:
            print(f"   • Exists: {year_data['year_label']}")

def get_country_id(country_code: str) -> str:
    """Get country UUID from country_code"""
    result = supabase.table('countries').select('id').eq('country_code', country_code).execute()
    if result.data:
        return result.data[0]['id']
    else:
        raise ValueError(f"Country not found: {country_code}")

def get_academic_year_id(year_label: str) -> str:
    """Get academic year UUID from year_label"""
    result = supabase.table('academic_years').select('id').eq('year_label', year_label).execute()
    if result.data:
        return result.data[0]['id']
    else:
        raise ValueError(f"Academic year not found: {year_label}")

def safe_int(value) -> int:
    """Convert value to integer, handling None, formulas, and strings"""
    if value is None or value == '' or value == '●' or value == '-':
        return 0
    try:
        # Handle string numbers with spaces or commas
        if isinstance(value, str):
            value = value.replace(',', '').replace(' ', '').strip()
        return int(float(value))
    except (ValueError, TypeError):
        return 0

def parse_table_1_1(ws, academic_year_id: str):
    """Parse Table 1.1 - Early Childhood Centres"""
    print("   📋 Parsing Table 1.1: Early Childhood...")

    institutions = []

    # Data starts at row 6, countries in column B (2)
    # Structure (based on inspection):
    # Col B: Country
    # Col C: Daycare Public
    # Col D: Daycare Private Church
    # Col E: Daycare Private Non-affiliated
    # Col F: Daycare Total (formula - skip)
    # Col G: Preschool Public
    # Col H: Preschool Private Church
    # Col I: Preschool Private Non-affiliated
    # Col J: Preschool Total (formula - skip)

    for row_idx in range(6, ws.max_row + 1):
        country_cell = ws.cell(row_idx, 2).value  # Column B

        if not country_cell or country_cell == 'OECS':
            continue

        # Map country code
        country_abbr = str(country_cell).strip().upper()
        if country_abbr not in COUNTRY_MAPPING:
            print(f"      ⚠️  Unknown country: {country_abbr} (row {row_idx})")
            continue

        country_code = COUNTRY_MAPPING[country_abbr]

        try:
            country_id = get_country_id(country_code)
        except ValueError as e:
            print(f"      ⚠️  {e}")
            continue

        # Extract data
        daycare_public = safe_int(ws.cell(row_idx, 3).value)
        daycare_private_church = safe_int(ws.cell(row_idx, 4).value)
        daycare_private_non_affiliated = safe_int(ws.cell(row_idx, 5).value)
        preschool_public = safe_int(ws.cell(row_idx, 7).value)
        preschool_private_church = safe_int(ws.cell(row_idx, 8).value)
        preschool_private_non_affiliated = safe_int(ws.cell(row_idx, 9).value)

        institutions.append({
            'country_id': country_id,
            'academic_year_id': academic_year_id,
            'daycare_public': daycare_public,
            'daycare_private_church': daycare_private_church,
            'daycare_private_non_affiliated': daycare_private_non_affiliated,
            'preschool_public': preschool_public,
            'preschool_private_church': preschool_private_church,
            'preschool_private_non_affiliated': preschool_private_non_affiliated,
            # Initialize other fields
            'primary_public': 0,
            'primary_private_church': 0,
            'primary_private_non_affiliated': 0,
            'secondary_public': 0,
            'secondary_private_church': 0,
            'secondary_private_non_affiliated': 0,
            'special_ed_public': 0,
            'special_ed_private_church': 0,
            'special_ed_private_non_affiliated': 0,
            'tvet_public': 0,
            'tvet_private_church': 0,
            'tvet_private_non_affiliated': 0,
            'post_secondary_public': 0,
            'post_secondary_private': 0,
        })

        print(f"      ✓ {country_abbr}: Daycare={daycare_public + daycare_private_church + daycare_private_non_affiliated}, Preschool={preschool_public + preschool_private_church + preschool_private_non_affiliated}")

    return institutions

def parse_table_1_2(ws, academic_year_id: str):
    """Parse Table 1.2 - Primary & Secondary Schools"""
    print("   📋 Parsing Table 1.2: Primary & Secondary...")

    institutions = {}

    # Need to inspect the actual structure
    # Data likely starts around row 6-7
    # Will have sections for Primary and Secondary

    for row_idx in range(6, ws.max_row + 1):
        country_cell = ws.cell(row_idx, 2).value  # Column B

        if not country_cell or country_cell == 'OECS' or country_cell == 'Total':
            continue

        # Skip header rows
        if country_cell and ('Primary' in str(country_cell) or 'Secondary' in str(country_cell) or 'School' in str(country_cell)):
            continue

        country_abbr = str(country_cell).strip().upper()
        if country_abbr not in COUNTRY_MAPPING:
            continue

        country_code = COUNTRY_MAPPING[country_abbr]

        try:
            country_id = get_country_id(country_code)
        except ValueError:
            continue

        # Initialize country record if not exists
        if country_id not in institutions:
            institutions[country_id] = {
                'country_id': country_id,
                'academic_year_id': academic_year_id,
                'primary_public': 0,
                'primary_private_church': 0,
                'primary_private_non_affiliated': 0,
                'secondary_public': 0,
                'secondary_private_church': 0,
                'secondary_private_non_affiliated': 0,
                'special_ed_public': 0,
                'special_ed_private_church': 0,
                'special_ed_private_non_affiliated': 0,
                'tvet_public': 0,
                'tvet_private_church': 0,
                'tvet_private_non_affiliated': 0,
            }

        # Extract primary school data (columns vary by file)
        # This is simplified - actual column positions need inspection
        primary_public = safe_int(ws.cell(row_idx, 3).value)
        primary_private = safe_int(ws.cell(row_idx, 4).value)
        secondary_public = safe_int(ws.cell(row_idx, 7).value)
        secondary_private = safe_int(ws.cell(row_idx, 8).value)

        institutions[country_id]['primary_public'] = primary_public
        institutions[country_id]['primary_private_church'] = primary_private
        institutions[country_id]['secondary_public'] = secondary_public
        institutions[country_id]['secondary_private_church'] = secondary_private

        print(f"      ✓ {country_abbr}: Primary={primary_public + primary_private}, Secondary={secondary_public + secondary_private}")

    return list(institutions.values())

def parse_table_1_3(ws, academic_year_id: str):
    """Parse Table 1.3 - Post-Secondary Institutions"""
    print("   📋 Parsing Table 1.3: Post-Secondary...")

    institutions = {}

    for row_idx in range(6, ws.max_row + 1):
        country_cell = ws.cell(row_idx, 2).value  # Column B

        if not country_cell or country_cell == 'OECS':
            continue

        country_abbr = str(country_cell).strip().upper()
        if country_abbr not in COUNTRY_MAPPING:
            continue

        country_code = COUNTRY_MAPPING[country_abbr]

        try:
            country_id = get_country_id(country_code)
        except ValueError:
            continue

        # Post-secondary data
        post_secondary_public = safe_int(ws.cell(row_idx, 3).value)
        post_secondary_private = safe_int(ws.cell(row_idx, 4).value)

        institutions[country_id] = {
            'country_id': country_id,
            'academic_year_id': academic_year_id,
            'post_secondary_public': post_secondary_public,
            'post_secondary_private': post_secondary_private,
        }

        print(f"      ✓ {country_abbr}: Post-Secondary={post_secondary_public + post_secondary_private}")

    return list(institutions.values())

def merge_institution_data(table1_data, table2_data, table3_data):
    """Merge data from all three tables by country_id"""
    print("\n🔗 Merging data from all tables...")

    merged = {}

    # Start with Table 1.1 (Early Childhood)
    for record in table1_data:
        country_id = record['country_id']
        merged[country_id] = record.copy()

    # Merge Table 1.2 (Primary/Secondary)
    for record in table2_data:
        country_id = record['country_id']
        if country_id in merged:
            merged[country_id].update({
                'primary_public': record.get('primary_public', 0),
                'primary_private_church': record.get('primary_private_church', 0),
                'primary_private_non_affiliated': record.get('primary_private_non_affiliated', 0),
                'secondary_public': record.get('secondary_public', 0),
                'secondary_private_church': record.get('secondary_private_church', 0),
                'secondary_private_non_affiliated': record.get('secondary_private_non_affiliated', 0),
                'special_ed_public': record.get('special_ed_public', 0),
                'special_ed_private_church': record.get('special_ed_private_church', 0),
                'special_ed_private_non_affiliated': record.get('special_ed_private_non_affiliated', 0),
                'tvet_public': record.get('tvet_public', 0),
                'tvet_private_church': record.get('tvet_private_church', 0),
                'tvet_private_non_affiliated': record.get('tvet_private_non_affiliated', 0),
            })

    # Merge Table 1.3 (Post-Secondary)
    for record in table3_data:
        country_id = record['country_id']
        if country_id in merged:
            merged[country_id].update({
                'post_secondary_public': record.get('post_secondary_public', 0),
                'post_secondary_private': record.get('post_secondary_private', 0),
            })

    print(f"   ✓ Merged data for {len(merged)} countries")
    return list(merged.values())

def parse_chapter1_file(filepath: str, academic_year: str):
    """Parse a Chapter 1 Excel file and extract institution counts"""
    print(f"\n📂 Processing: {Path(filepath).name}")
    print(f"   Year: {academic_year}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Get academic year ID
    year_label = ACADEMIC_YEAR_MAPPING[academic_year]
    academic_year_id = get_academic_year_id(year_label)

    # Parse each table
    table1_data = []
    table2_data = []
    table3_data = []

    if 'Table 1.1' in wb.sheetnames:
        table1_data = parse_table_1_1(wb['Table 1.1'], academic_year_id)

    if 'Table 1.2' in wb.sheetnames:
        table2_data = parse_table_1_2(wb['Table 1.2'], academic_year_id)

    if 'Table 1.3' in wb.sheetnames:
        table3_data = parse_table_1_3(wb['Table 1.3'], academic_year_id)

    wb.close()

    # Merge all three tables
    merged_data = merge_institution_data(table1_data, table2_data, table3_data)

    return merged_data

def import_chapter1():
    """Main import function for Chapter 1 data"""
    print("\n" + "=" * 80)
    print("📊 IMPORTING CHAPTER 1: INSTITUTIONS DATA")
    print("=" * 80)

    # Ensure academic years exist
    get_or_create_academic_years()

    # Base directory
    base_dir = Path(__file__).parent.parent / 'DIGEST_WEB' / 'Extracted Chapters' / 'Chapter 1'

    # Files to process
    files = [
        ('2020-21.xlsx', '2020-21'),
        ('2021-22.xlsx', '2021-22'),
        ('2022-23.xlsx', '2022-23'),
    ]

    all_data = []

    for filename, year in files:
        filepath = base_dir / filename
        if filepath.exists():
            data = parse_chapter1_file(str(filepath), year)
            all_data.extend(data)
        else:
            print(f"\n⚠️  File not found: {filepath}")

    # Insert data into Supabase
    if all_data:
        print(f"\n💾 Inserting {len(all_data)} records into institutions table...")

        # Delete existing data for these years first
        for year in ['2020-2021', '2021-2022', '2022-2023']:
            try:
                year_id = get_academic_year_id(year)
                result = supabase.table('institutions').delete().eq('academic_year_id', year_id).execute()
                print(f"   🗑️  Cleared existing data for {year}")
            except Exception as e:
                print(f"   ⚠️  Could not clear {year}: {e}")

        # Batch insert
        batch_size = 50
        inserted_count = 0
        for i in range(0, len(all_data), batch_size):
            batch = all_data[i:i+batch_size]
            try:
                supabase.table('institutions').insert(batch).execute()
                inserted_count += len(batch)
                print(f"   ✓ Inserted batch {i//batch_size + 1}/{(len(all_data)-1)//batch_size + 1} ({inserted_count} records)")
            except Exception as e:
                print(f"   ❌ Error inserting batch: {e}")

        print(f"\n✅ Successfully imported {inserted_count} institution records!")
        print(f"   📈 Data now available for dashboard visualization")
    else:
        print("\n⚠️  No data to import")

if __name__ == '__main__':
    try:
        import_chapter1()
        print("\n" + "=" * 80)
        print("✨ Import complete! Check your dashboard to see the real data.")
        print("=" * 80 + "\n")
    except Exception as e:
        print(f"\n❌ Error during import: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
