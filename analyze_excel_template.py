import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import json

def analyze_worksheet(ws):
    """Comprehensive analysis of a worksheet"""
    analysis = {
        'name': ws.title,
        'dimensions': str(ws.dimensions),
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'merged_cells': [],
        'headers': [],
        'data_structure': [],
        'formulas': [],
        'data_validations': [],
        'conditional_formatting': [],
        'sample_data': [],
        'column_details': {}
    }

    # Get merged cells
    for merged_range in ws.merged_cells.ranges:
        analysis['merged_cells'].append(str(merged_range))

    # Analyze first 20 rows to understand structure
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)

            # Skip if part of merged cell
            if isinstance(cell, MergedCell):
                row_data.append({'value': '[MERGED]', 'col': get_column_letter(col_idx)})
                continue

            cell_info = {
                'col': get_column_letter(col_idx),
                'value': cell.value,
                'data_type': cell.data_type,
                'number_format': cell.number_format,
            }

            # Check for formulas
            if cell.data_type == 'f':
                cell_info['formula'] = cell.value
                analysis['formulas'].append({
                    'cell': f"{get_column_letter(col_idx)}{row_idx}",
                    'formula': cell.value
                })

            # Check for hyperlinks
            if cell.hyperlink:
                cell_info['hyperlink'] = str(cell.hyperlink.target)

            # Check for comments
            if cell.comment:
                cell_info['comment'] = cell.comment.text

            row_data.append(cell_info)

        analysis['data_structure'].append({
            'row': row_idx,
            'cells': row_data
        })

    # Get data validations
    for dv in ws.data_validations.dataValidation:
        validation_info = {
            'type': dv.type,
            'formula1': dv.formula1,
            'formula2': dv.formula2,
            'sqref': str(dv.sqref),
            'allow_blank': dv.allowBlank,
            'show_dropdown': dv.showDropDown,
            'prompt': dv.prompt,
            'prompt_title': dv.promptTitle,
            'error': dv.error,
            'error_title': dv.errorTitle
        }
        analysis['data_validations'].append(validation_info)

    # Analyze column patterns (look at more rows for data)
    for col_idx in range(1, min(ws.max_column + 1, 50)):
        col_letter = get_column_letter(col_idx)
        col_values = []
        col_types = set()
        has_formula = False

        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row_idx, col_idx)
            if not isinstance(cell, MergedCell) and cell.value is not None:
                col_values.append(str(cell.value)[:100])  # Limit length
                col_types.add(cell.data_type)
                if cell.data_type == 'f':
                    has_formula = True

        if col_values:
            analysis['column_details'][col_letter] = {
                'sample_values': col_values[:10],
                'data_types': list(col_types),
                'has_formula': has_formula,
                'non_empty_count': len(col_values)
            }

    return analysis

def main():
    file_path = r"C:\Users\Clendon\oecs-education-dashboard\DIGEST_WEB\Blank OECS MS Template.xlsx"

    print("="*80)
    print("COMPREHENSIVE EXCEL TEMPLATE ANALYSIS")
    print("="*80)
    print(f"\nFile: {file_path}\n")

    wb = openpyxl.load_workbook(file_path, data_only=False)

    print(f"Total worksheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}\n")

    all_analyses = {}

    for sheet_name in wb.sheetnames:
        print("\n" + "="*80)
        print(f"ANALYZING SHEET: {sheet_name}")
        print("="*80)

        ws = wb[sheet_name]
        analysis = analyze_worksheet(ws)
        all_analyses[sheet_name] = analysis

        print(f"\nDimensions: {analysis['dimensions']}")
        print(f"Max Row: {analysis['max_row']}, Max Column: {analysis['max_col']}")
        print(f"Merged Cells: {len(analysis['merged_cells'])}")

        if analysis['merged_cells']:
            print("\nMerged Cell Ranges:")
            for merged in analysis['merged_cells'][:10]:
                print(f"  - {merged}")
            if len(analysis['merged_cells']) > 10:
                print(f"  ... and {len(analysis['merged_cells']) - 10} more")

        print(f"\nData Validations: {len(analysis['data_validations'])}")
        if analysis['data_validations']:
            for i, dv in enumerate(analysis['data_validations'][:5]):
                print(f"\n  Validation {i+1}:")
                print(f"    Type: {dv['type']}")
                print(f"    Range: {dv['sqref']}")
                print(f"    Formula1: {dv['formula1']}")
                if dv['prompt']:
                    print(f"    Prompt: {dv['prompt']}")
            if len(analysis['data_validations']) > 5:
                print(f"\n  ... and {len(analysis['data_validations']) - 5} more validations")

        print(f"\nFormulas Found: {len(analysis['formulas'])}")
        if analysis['formulas']:
            print("\nSample Formulas:")
            for formula in analysis['formulas'][:10]:
                print(f"  {formula['cell']}: {formula['formula']}")
            if len(analysis['formulas']) > 10:
                print(f"  ... and {len(analysis['formulas']) - 10} more formulas")

        print(f"\nColumns with Data: {len(analysis['column_details'])}")

        print("\n" + "-"*80)
        print("FIRST 20 ROWS STRUCTURE:")
        print("-"*80)

        for row_info in analysis['data_structure']:
            row_num = row_info['row']
            non_empty = [c for c in row_info['cells'] if c['value'] not in [None, '[MERGED]']]

            if non_empty or row_num <= 10:  # Always show first 10 rows
                print(f"\nRow {row_num}:")
                for cell in row_info['cells']:
                    if cell['value'] not in [None, '[MERGED]']:
                        val_str = str(cell['value'])[:80]
                        print(f"  {cell['col']}: {val_str}")
                        if cell.get('formula'):
                            print(f"      [Formula: {cell['formula']}]")
                        if cell.get('comment'):
                            print(f"      [Comment: {cell['comment']}]")
                        if cell['number_format'] != 'General':
                            print(f"      [Format: {cell['number_format']}]")

    # Save detailed JSON report
    output_file = r"C:\Users\Clendon\oecs-education-dashboard\template_analysis_report.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_analyses, f, indent=2, ensure_ascii=False)

    print("\n" + "="*80)
    print(f"Detailed JSON report saved to: {output_file}")
    print("="*80)

if __name__ == "__main__":
    main()
